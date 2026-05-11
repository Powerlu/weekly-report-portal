# -*- coding: utf-8 -*-
import json
import glob
import os

# Find Excel file
desktop = r'C:\Users\13328\Desktop'
files = glob.glob(desktop + r'\*重点工作任务监控表*.xlsx')
if not files:
    print("ERROR: Excel file not found on desktop")
    exit(1)

excel_path = files[0]
print(f"Reading: {excel_path}")

import openpyxl
wb = openpyxl.load_workbook(excel_path, data_only=True)

all_tasks = []

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    print(f"\nSheet: {sheet_name}")
    
    rows = list(ws.iter_rows(values_only=True))
    
    # Find header row
    header_idx = None
    for i, row in enumerate(rows):
        if row and '任务名称' in str(row[0]):
            header_idx = i
            break
    if header_idx is None:
        print(f"  Skipping - no header found")
        continue
    
    current_project = ''
    current_modules = []
    current_status = ''
    current_achievements = ''
    current_blockers = ''
    current_support = ''
    current_assignee = ''
    current_target_date = ''
    
    def flush_task():
        if not current_modules:
            return
        task_name = '、'.join(m for m in current_modules if m)
        if not task_name:
            return
        
        def clean(s):
            s = s.replace('\n', ' ').replace('\r', ' ')
            while '  ' in s:
                s = s.replace('  ', ' ')
            return s.strip()
        
        status_map = {
            '已完成': '已完成', '完成': '已完成',
            '进行中': '进行中', '未完成': '进行中',
            '持续跟进': '进行中', '延期': '延期', '阻塞': '阻塞',
        }
        status = status_map.get(current_status, current_status) if current_status else '进行中'
        
        all_tasks.append({
            "project_name": current_project or sheet_name,
            "task_name": clean(task_name)[:150],
            "module": "",
            "target_date": clean(current_target_date),
            "assignee": clean(current_assignee),
            "status": status,
            "achievements": clean(current_achievements),
            "blockers": clean(current_blockers),
            "support_needed": clean(current_support)
        })
    
    for i in range(header_idx + 2, len(rows)):
        row = rows[i]
        if not row or all(not str(c).strip() for c in row if c):
            continue
        
        col0 = str(row[0]).strip() if row[0] else ''
        col1 = str(row[1]).strip() if len(row) > 1 and row[1] else ''
        col2 = str(row[2]).strip() if len(row) > 2 and row[2] else ''
        col5 = str(row[5]).strip() if len(row) > 5 and row[5] else ''
        col6 = str(row[6]).strip() if len(row) > 6 and row[6] else ''
        col7 = str(row[7]).strip() if len(row) > 7 and row[7] else ''
        col8 = str(row[8]).strip() if len(row) > 8 and row[8] else ''
        col9 = str(row[9]).strip() if len(row) > 9 and row[9] else ''
        
        if col0:
            flush_task()
            current_project = col0
            current_modules = []
            current_status = col5
            current_achievements = col6
            current_blockers = col7
            current_support = col8
            current_assignee = col9
            current_target_date = col2
            if col1:
                current_modules.append(col1.split('\n')[0])
        elif col1 and col5:
            if col6 or col5 != current_status:
                flush_task()
                current_modules = [col1.split('\n')[0]]
                current_status = col5
                current_achievements = col6
                current_blockers = col7
                current_support = col8
                current_assignee = col9 or current_assignee
                current_target_date = col2 or current_target_date
            else:
                current_modules.append(col1.split('\n')[0])
    
    flush_task()

# ===== 过滤掉王荣欣 =====
before = len(all_tasks)
all_tasks = [t for t in all_tasks if t['assignee'] != '王荣欣']
after = len(all_tasks)
print(f"\n过滤王荣欣: {before} → {after} 条")

# Summary
from collections import Counter
project_counts = Counter(t['project_name'] for t in all_tasks)
print(f"\n最终共 {len(all_tasks)} 条任务:")
for proj, cnt in project_counts.items():
    print(f"  {proj}: {cnt} 条")

# Write
out_path = r'C:\Users\13328\weekly-report-portal\src\assets\data.json'
with open(out_path, 'w', encoding='utf-8') as f:
    json.dump(all_tasks, f, ensure_ascii=False, indent=2)

print(f"\n已写入: {out_path}")
